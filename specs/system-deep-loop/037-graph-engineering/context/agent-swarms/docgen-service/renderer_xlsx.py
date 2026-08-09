"""
Server-side Excel renderer (openpyxl) + LibreOffice recalc.

Consumes a MATERIALIZED plan (the client already ran each data sheet's SQL over
the user's real rows, so every sheet is literal here):
  { "sheets": [ { "name": str, "headers": [str],
                  "rows": [[ cell ]] } ] }
where a cell is a literal (str/number/bool/null) or a live formula
  { "formula": "SUM(B2:B10)", "format"?: "$#,##0.00" }   # A1 syntax, no leading "="

Applies the Anthropic xlsx-skill guidance:
  • formulas written as strings, then recalculated by LibreOffice so the file
    opens showing cached VALUES (not blank cells) in any viewer,
  • number formats applied per formula cell,
  • bold shaded header row, frozen header, sensible auto column widths,
  • sheet names sanitised to Excel's constraints.
"""
from __future__ import annotations

import io
import os
import re
import subprocess
import tempfile
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="EEF2F7")
HEADER_FONT = Font(bold=True, color="1F2937")
THIN = Side(style="thin", color="D9DEE5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
_BAD_SHEET = re.compile(r"[\[\]:*?/\\]")


def _sheet_name(raw: str, used: set[str]) -> str:
    name = _BAD_SHEET.sub(" ", str(raw or "Sheet")).strip()[:31] or "Sheet"
    base, i = name, 2
    while name.lower() in used:
        suffix = f" ({i})"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(name.lower())
    return name


def _write_cell(ws, r: int, c: int, value: Any) -> int:
    """Write one cell; return its display length for column-width sizing.

    A cell with no value is not written AT ALL — not even for its border.

    openpyxl was previously asked to materialise every blank as a styled,
    value-less cell (`<c r="B4" s="5"/>`). Written that way the file is
    correct, but the LibreOffice recalc round-trip below then collapses a run
    of value-less cells and drags the next formula left into the gap. Measured
    on a totals row `["Total", None x6, SUM(H2:H3), SUM(I2:I3)]`:

        openpyxl only          H4 =SUM(H2:H3)     correct
        after LibreOffice      B4 =SUM(H2:H3)     moved to the first blank

    which lands a column's total under an unrelated heading and empties the
    cell every roll-up sheet references. Padding with "" behaves identically;
    padding with 0 keeps the position, which is what identified value-less
    cells as the trigger. Omitting them entirely is the fix, at the cost of no
    grid border on genuinely empty cells.
    """
    if value is None:
        return 0
    cell = ws.cell(row=r, column=c)
    cell.border = BORDER
    if isinstance(value, dict) and "formula" in value:
        formula = str(value["formula"]).lstrip("=")
        cell.value = "=" + formula
        fmt = value.get("format")
        if fmt:
            cell.number_format = str(fmt)
        return min(len(formula) + 2, 24)
    if isinstance(value, bool):
        cell.value = value
        return 5
    if isinstance(value, (int, float)):
        cell.value = value
        # A light default numeric format keeps big numbers readable.
        if isinstance(value, float) or abs(value) >= 1000:
            cell.number_format = "#,##0.##"
        return len(str(value)) + 1
    cell.value = str(value)
    return len(str(value))


# LibreOffice does NOT recalculate foreign-format (xlsx) formulas on load unless
# told to — so a plain `--convert-to` round-trip saves the formulas with no
# cached value and viewers that don't recalc show blank cells. This profile
# fragment forces "Always recalculate" for both ODF and OOXML on load.
_RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>
"""


def _recalc(xlsx_bytes: bytes) -> bytes:
    """Round-trip through LibreOffice so formula results are cached in the file.

    With the recalc-on-load profile below, LibreOffice recomputes every formula
    when it opens the workbook and stores the values on save, so
    Excel/Sheets/Numbers show numbers immediately instead of blank formula
    cells. Best-effort: returns the input unchanged if soffice is unavailable.
    """
    with tempfile.TemporaryDirectory() as d:
        src = os.path.join(d, "book.xlsx")
        with open(src, "wb") as f:
            f.write(xlsx_bytes)
        # Convert into a SEPARATE dir — LibreOffice silently refuses to overwrite
        # the source file it is reading, so outdir must not equal the src's dir.
        out_dir = os.path.join(d, "out")
        os.makedirs(out_dir, exist_ok=True)
        # Pre-seed the profile with the "always recalc on load" setting.
        profile = os.path.join(d, "lo")
        user_dir = os.path.join(profile, "user")
        os.makedirs(user_dir, exist_ok=True)
        with open(os.path.join(user_dir, "registrymodifications.xcu"), "w") as f:
            f.write(_RECALC_XCU)
        try:
            subprocess.run(
                ["soffice", f"-env:UserInstallation=file://{profile}",
                 "--headless", "--calc", "--convert-to",
                 "xlsx:Calc MS Excel 2007 XML", "--outdir", out_dir, src],
                check=True, timeout=120, capture_output=True,
            )
        except Exception:
            return xlsx_bytes
        out = os.path.join(out_dir, "book.xlsx")
        if not os.path.exists(out):
            return xlsx_bytes
        with open(out, "rb") as f:
            return f.read()


def render_xlsx(plan: dict[str, Any], recalc: bool = True) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()

    sheets = plan.get("sheets") or []
    if not sheets:
        sheets = [{"name": "Sheet1", "headers": ["Note"], "rows": [["No data"]]}]

    for sheet in sheets:
        if not isinstance(sheet, dict):
            continue
        ws = wb.create_sheet(_sheet_name(sheet.get("name", "Sheet"), used))
        headers = [str(h) for h in (sheet.get("headers") or [])]
        rows = sheet.get("rows") or []
        ncols = max(len(headers), max((len(r) for r in rows if isinstance(r, list)), default=0), 1)
        widths = [len(headers[i]) if i < len(headers) else 0 for i in range(ncols)]

        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c, value=headers[c - 1] if c - 1 < len(headers) else "")
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

        for ri, row in enumerate(rows, start=2):
            if not isinstance(row, list):
                continue
            for ci in range(1, ncols + 1):
                val = row[ci - 1] if ci - 1 < len(row) else None
                w = _write_cell(ws, ri, ci, val)
                widths[ci - 1] = max(widths[ci - 1], w)

        for i in range(ncols):
            ws.column_dimensions[get_column_letter(i + 1)].width = min(max(widths[i] + 2, 8), 60)
        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    return _recalc(data) if recalc else data
